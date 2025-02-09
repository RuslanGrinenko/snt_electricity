import os
import csv
import requests
from flask import Flask, render_template, request, send_file, redirect
from openpyxl import Workbook, load_workbook
from datetime import datetime, timezone

app = Flask(__name__)
app.secret_key = 'supersecretkey'

def get_api_key():
    with open('api.key', 'r') as f:
        return f.read().strip()

def read_recipies():
    recipies = []
    with open('recipies.csv', 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            # Добавляем поля по умолчанию
            default_row = {
                'T1_FROM': 0.0,
                'T1_TO': 0.0,
                'T1_CONS': 0.0,
                'T2_FROM': 0.0,
                'T2_TO': 0.0,
                'T2_CONS': 0.0,
                'SUMM': 0.0,
                'SUMM_LOSS': 0.0
            }
            default_row.update(row)
            recipies.append(default_row)
    return recipies

def fetch_data(modem_id, channel, from_date, to_date):
    api_key = get_api_key()
    url = f"https://lk.waviot.ru/api.data/get_modem_channel_values/?channel=electro_ac_p_lsum_{channel}&key={api_key}&modem_id={modem_id}&from={from_date}&to={to_date}"

    try:
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        data = response.json()
    except Exception as e:
        print(f"API Error: {str(e)}")
        return None, None

    if data.get('status') != 'ok':
        print(f"API Status Error: {data.get('message', 'Unknown error')}")
        return None, None

    values = data.get('values', {})
    sorted_ts = sorted([int(ts) for ts in values.keys()])

    from_val = next((ts for ts in sorted_ts if ts >= from_date), None)
    to_val = next((ts for ts in reversed(sorted_ts) if ts <= to_date), None)

    from_value = values.get(str(from_val)) if from_val else None
    to_value = values.get(str(to_val)) if to_val else None

    return from_value, to_value

def process_data(recipies, start_date, end_date, tax_day, tax_night):
    start_ts = int(start_date.timestamp())
    end_ts = int(end_date.timestamp())

    general_t1_cons = 0
    general_t2_cons = 0
    total_t1_cons = 0
    total_t2_cons = 0

    for counter in recipies:
        modem_id = counter['MODEM_ID']

        multiplier = 50 if counter['LOC'] == 'Общий' else 1
        t1_from, t1_to = fetch_data(modem_id, 't1', start_ts, end_ts)
        t1_from = float(t1_from*multiplier) if t1_from is not None else 0.0
        t1_to = float(t1_to*multiplier) if t1_to is not None else 0.0

        t2_from, t2_to = fetch_data(modem_id, 't2', start_ts, end_ts)
        t2_from = float(t2_from*multiplier) if t2_from is not None else 0.0
        t2_to = float(t2_to*multiplier) if t2_to is not None else 0.0


        t1_cons = t1_to - t1_from
        t2_cons = t2_to - t2_from

        counter.update({
            'T1_FROM': t1_from,
            'T1_TO': t1_to,
            'T1_CONS': t1_cons,
            'T2_FROM': t2_from,
            'T2_TO': t2_to,
            'T2_CONS': t2_cons,
            'SUMM': t1_cons * tax_day + t2_cons * tax_night
        })

        if counter['LOC'] == 'Общий':
            general_t1_cons = t1_cons
            general_t2_cons = t2_cons
        else:
            total_t1_cons += t1_cons
            total_t2_cons += t2_cons

    loss_day = (general_t1_cons - total_t1_cons)*100/total_t1_cons if general_t1_cons else 0
    loss_night = (general_t2_cons - total_t2_cons)*100/total_t2_cons if general_t2_cons else 0

    for counter in recipies:
        if counter['LOC'] != 'Общий':
            t1_cons = counter['T1_CONS']
            t2_cons = counter['T2_CONS']
            summ_loss = (t1_cons + t1_cons * loss_day/100) * tax_day + \
                       (t2_cons + t2_cons * loss_night/100) * tax_night
            counter['SUMM_LOSS'] = summ_loss
        else:
            counter['SUMM_LOSS'] = 0

    return loss_day, loss_night

@app.route('/', methods=['GET', 'POST'])
def index():
    recipies = read_recipies()
    loss_day = loss_night = 0
    default_tax_day = 6.28
    default_tax_night = 3.4

    if request.method == 'POST':
        tax_day = request.form.get('tax_day', '').strip()
        tax_day = float(tax_day) if tax_day else default_tax_day

        tax_night = request.form.get('tax_night', '').strip()
        tax_night = float(tax_night) if tax_night else default_tax_night

        start_date = datetime.strptime(request.form['start_date'], '%Y-%m-%d')
        end_date = datetime.strptime(request.form['end_date'], '%Y-%m-%d')

        loss_day, loss_night = process_data(
            recipies, start_date, end_date, tax_day, tax_night
        )

        return render_template('index.html',
                             recipies=recipies,
                             tax_day=tax_day,
                             tax_night=tax_night,
                             start_date=start_date.strftime('%Y-%m-%d'),
                             end_date=end_date.strftime('%Y-%m-%d'),
                             loss_day=loss_day,
                             loss_night=loss_night)

    return render_template('index.html',
                         recipies=recipies,
                         tax_day=default_tax_day,
                         tax_night=default_tax_night)

@app.route('/save', methods=['POST'])
def save_file():
    try:
        data = request.form.to_dict()

        # Получаем порядок modem_id из скрытого поля
        order = data.get('order', '').split(',') if data.get('order') else []

        # Собираем данные в правильном порядке
        recipies = []
        for modem_id in order:
            if not modem_id:
                continue

            recipies.append({
                'LOC': data.get(f'loc_{modem_id}', ''),
                'MODEM_ID': modem_id,
                'E_MAIL': data.get(f'email_{modem_id}', ''),
                'T1_FROM': float(data.get(f't1_from_{modem_id}', 0)),
                'T1_TO': float(data.get(f't1_to_{modem_id}', 0)),
                'T1_CONS': float(data.get(f't1_cons_{modem_id}', 0)),
                'T2_FROM': float(data.get(f't2_from_{modem_id}', 0)),
                'T2_TO': float(data.get(f't2_to_{modem_id}', 0)),
                'T2_CONS': float(data.get(f't2_cons_{modem_id}', 0)),
                'SUMM': float(data.get(f'summ_{modem_id}', 0)),
                'SUMM_LOSS': float(data.get(f'summ_loss_{modem_id}', 0))
            })

        # Создаем XLSX
        wb = Workbook()
        ws = wb.active

        # Метаданные
        ws.append(['Start Date', data.get('start_date', '')])
        ws.append(['End Date', data.get('end_date', '')])
        ws.append(['Tax Day', data.get('tax_day', '')])
        ws.append(['Tax Night', data.get('tax_night', '')])
        ws.append(['Loss Day (%)', data.get('loss_day', '0')])
        ws.append(['Loss Night (%)', data.get('loss_night', '0')])

        # Заголовки
        headers = ['LOC', 'MODEM_ID', 'E_MAIL', 'T1_FROM', 'T1_TO', 'T1_CONS',
                 'T2_FROM', 'T2_TO', 'T2_CONS', 'SUMM', 'SUMM_LOSS']
        ws.append(headers)

        # Данные в правильном порядке
        for counter in recipies:
            ws.append([
                counter['LOC'],
                counter['MODEM_ID'],
                counter['E_MAIL'],
                counter['T1_FROM'],
                counter['T1_TO'],
                counter['T1_CONS'],
                counter['T2_FROM'],
                counter['T2_TO'],
                counter['T2_CONS'],
                counter['SUMM'],
                counter['SUMM_LOSS']
            ])

        filename = 'current_data.xlsx'
        wb.save(filename)

        return send_file(filename, as_attachment=True)

    except Exception as e:
        return f"Ошибка сохранения: {str(e)}", 500

@app.route('/load', methods=['POST'])
def load_file():
    if 'file' not in request.files:
        return redirect('/')

    file = request.files['file']
    if file.filename == '':
        return redirect('/')

    try:
        wb = load_workbook(file)
        ws = wb.active

        # Чтение метаданных с потерями
        start_date = ws['B1'].value
        end_date = ws['B2'].value
        tax_day = float(ws['B3'].value) if ws['B3'].value else 6.28
        tax_night = float(ws['B4'].value) if ws['B4'].value else 3.4
        loss_day = float(ws['B5'].value) if ws['B5'].value else 0.0  # Новая строка
        loss_night = float(ws['B6'].value) if ws['B6'].value else 0.0  # Новая строка

        headers = [cell.value for cell in ws[7]]  # Смещаем из-за новых строк
        recipies = []

        for row in ws.iter_rows(min_row=8, values_only=True):  # Смещаем начало данных
            if not any(row):
                continue
            data = dict(zip(headers, row))
            for field in ['T1_FROM', 'T1_TO', 'T1_CONS',
                        'T2_FROM', 'T2_TO', 'T2_CONS',
                        'SUMM', 'SUMM_LOSS']:
                data[field] = float(data[field]) if data.get(field) else 0.0
            recipies.append(data)

        return render_template('index.html',
                            recipies=recipies,
                            tax_day=tax_day,
                            tax_night=tax_night,
                            start_date=start_date,
                            end_date=end_date,
                            loss_day=loss_day,  # Передаём потери
                            loss_night=loss_night)

    except Exception as e:
        return f"Ошибка загрузки файла: {str(e)}", 500

if __name__ == '__main__':
    app.run(debug=True)
